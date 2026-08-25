import os
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from config import settings
from sqlalchemy.orm import Session
from models import LandRecord


def generate_excel_export(db: Session, record_ids: list[int] = None) -> str:
    """
    Generates a multi-sheet Excel export for land records, validations, and audit logs.
    """
    wb = Workbook()
    
    # Build query
    query = db.query(LandRecord)
    if record_ids:
        query = query.filter(LandRecord.id.in_(record_ids))
    records = query.all()

    # --- Sheet 1: Land Records ---
    ws_records = wb.active
    ws_records.title = "Land Records"
    
    headers = [
        "Record ID", "Document ID", "Status", "Owner Name", "Survey Number", 
        "Village", "Taluk/Tehsil", "District", "Area", "Area Unit", "Land Type",
        "Registration No", "Registration Date", "Mutation No", "Mutation Date",
        "AI Confidence", "Verified By", "Verified At"
    ]
    
    _write_header(ws_records, headers)
    
    for r, row_idx in zip(records, range(2, len(records) + 2)):
        ws_records.cell(row=row_idx, column=1, value=r.id)
        ws_records.cell(row=row_idx, column=2, value=r.document_id)
        ws_records.cell(row=row_idx, column=3, value=r.status)
        ws_records.cell(row=row_idx, column=4, value=r.owner_name)
        ws_records.cell(row=row_idx, column=5, value=r.survey_number)
        ws_records.cell(row=row_idx, column=6, value=r.village)
        ws_records.cell(row=row_idx, column=7, value=r.taluk_tehsil)
        ws_records.cell(row=row_idx, column=8, value=r.district)
        ws_records.cell(row=row_idx, column=9, value=r.area)
        ws_records.cell(row=row_idx, column=10, value=r.area_unit)
        ws_records.cell(row=row_idx, column=11, value=r.land_type)
        ws_records.cell(row=row_idx, column=12, value=r.registration_number)
        ws_records.cell(row=row_idx, column=13, value=r.registration_date)
        ws_records.cell(row=row_idx, column=14, value=r.mutation_number)
        ws_records.cell(row=row_idx, column=15, value=r.mutation_date)
        
        conf = f"{round((r.overall_confidence or 0) * 100)}%"
        ws_records.cell(row=row_idx, column=16, value=conf)
        
        if r.verifier:
            ws_records.cell(row=row_idx, column=17, value=r.verifier.username)
        if r.verified_at:
            ws_records.cell(row=row_idx, column=18, value=r.verified_at.strftime("%Y-%m-%d %H:%M:%S"))

    # --- Sheet 2: Validation Issues ---
    ws_val = wb.create_sheet(title="Validation Issues")
    val_headers = ["Record ID", "Rule", "Field", "Severity", "Message", "Resolved"]
    _write_header(ws_val, val_headers)
    
    row_idx = 2
    for r in records:
        for val in r.validation_results:
            ws_val.cell(row=row_idx, column=1, value=r.id)
            ws_val.cell(row=row_idx, column=2, value=val.rule_name)
            ws_val.cell(row=row_idx, column=3, value=val.field_name)
            ws_val.cell(row=row_idx, column=4, value=val.severity)
            ws_val.cell(row=row_idx, column=5, value=val.message)
            ws_val.cell(row=row_idx, column=6, value="Yes" if val.is_resolved else "No")
            row_idx += 1

    # --- Sheet 3: Audit Logs ---
    ws_audit = wb.create_sheet(title="Audit History")
    audit_headers = ["Record ID", "Timestamp", "User", "Action", "Field", "Old Value", "New Value"]
    _write_header(ws_audit, audit_headers)
    
    row_idx = 2
    for r in records:
        for log in r.audit_logs:
            ws_audit.cell(row=row_idx, column=1, value=r.id)
            ws_audit.cell(row=row_idx, column=2, value=log.timestamp.strftime("%Y-%m-%d %H:%M:%S") if log.timestamp else "")
            ws_audit.cell(row=row_idx, column=3, value=log.user.username if log.user else "System")
            ws_audit.cell(row=row_idx, column=4, value=log.action)
            ws_audit.cell(row=row_idx, column=5, value=log.field_name)
            ws_audit.cell(row=row_idx, column=6, value=log.previous_value)
            ws_audit.cell(row=row_idx, column=7, value=log.new_value)
            row_idx += 1

    # Save
    filename = f"export_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(settings.GENERATED_DIR, filename)
    wb.save(filepath)
    
    return filepath


def _write_header(ws, headers):
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid") # Blue-500
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = font
        cell.fill = fill
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)
